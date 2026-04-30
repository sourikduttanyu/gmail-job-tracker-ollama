#!/usr/bin/env python3
"""
Job Tracker — scans Gmail for job application emails, saves to jobs.xlsx.
Re-run anytime to pull new emails and update the sheet.

Usage:
    python3 job_tracker.py               # regex-only classification
    python3 job_tracker.py --ml          # Ollama for Unknown emails (default: llama3.2)
    python3 job_tracker.py --ml --model phi3:mini
"""

import argparse
import base64
import json
import re
from datetime import datetime
from pathlib import Path

import requests as http
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build
import pandas as pd

# ── Config ──────────────────────────────────────────────────────────────────
SCOPES        = ['https://www.googleapis.com/auth/gmail.readonly']
BASE_DIR      = Path(__file__).parent
CREDS_FILE    = BASE_DIR / 'credentials.json'
TOKEN_FILE    = BASE_DIR / 'token.json'
OUTPUT_FILE   = BASE_DIR / 'jobs.xlsx'
MAX_RESULTS   = 500

OLLAMA_URL          = 'http://localhost:11434/api/chat'
OLLAMA_MODEL_DEFAULT = 'llama3.2'
VALID_STATUSES      = {'Offer', 'Rejected', 'Interview', 'Applied', 'Unknown'}

# ── Gmail search query ───────────────────────────────────────────────────────
GMAIL_QUERY = (
    'subject:(application OR interview OR offer OR rejection OR '
    '"thank you for applying" OR "application received" OR '
    '"coding challenge" OR assessment OR recruiter OR "phone screen" OR '
    '"we regret" OR unfortunately OR "next steps" OR "moving forward" OR '
    '"not selected" OR "other candidates") '
    'OR from:(linkedin.com OR indeed.com OR greenhouse.io OR lever.co OR '
    'myworkday.com OR glassdoor.com OR recruiting OR talent OR careers OR '
    'noreply OR no-reply)'
)

# ── Blocked senders / topics — never process these ───────────────────────────
# Matched against sender name + sender email + subject (case-insensitive)
BLOCKED_SENDER_PATTERNS = [
    r'act.?uaw',
    r'local 7902',
    r'new york yankees',
    r'desi party',
    r'kaggle',
    r'eventbrite',
    r'peoplesync',
    r'edward carr',
    r'the economist',
    r'nyu it.*survey',
    r'user satisfaction survey',
    r'tassel',
    r'mailsuite',
    r'nyu tandon career hub',
    r'glassdoor jobs',
    r'office of records',
    r'records and registration',
    r'peter voltz',
]


def is_blocked(sender: str, subject: str) -> bool:
    haystack = (sender + ' ' + subject).lower()
    return any(re.search(p, haystack) for p in BLOCKED_SENDER_PATTERNS)


# ── Academic email filter ────────────────────────────────────────────────────
ACADEMIC_CONTENT_PATTERNS = [
    r'\bcourse\b', r'\bcoursework\b', r'\bsyllabus\b',
    r'\blecture\b', r'\boffice hours?\b', r'\bprofessor\b', r'\bprof\b',
    r'\bregistrar\b', r'\bfinancial aid\b', r'\btuition\b', r'\bbursar\b',
    r'\btranscript\b', r'\bgrade[sd]?\b', r'\bGPA\b',
    r'\bsemester\b', r'\bquarter\b', r'\bacademic year\b',
    r'\bcommencement\b', r'\bgraduation ceremony\b',
    r'\bcampus\b', r'\bdining hall\b', r'\bdormitory\b', r'\bresidence hall\b',
    r'\bstudent (?:id|account|portal|services|union|government)\b',
    r'\blibrary\b', r'\bfaculty\b', r'\bcurriculum\b',
    r'\bTA\b', r'\bteaching assistant\b',
    r'\bhomework\b', r'\bmidterm\b', r'\bfinal exam\b',
    r'\bdean\b', r'\bdepartment of \w+\b',
    r'\benrollment\b', r'\badmission[s]?\b',
    r'\bstudent loan\b', r'\bfafsa\b',
    r'\bclub meeting\b', r'\bstudent org\b',
]

JOB_OVERRIDE_PATTERNS = [
    r'career(?:s| center| fair|\.)', r'recruiting', r'internship',
    r'full.?time offer', r'job fair', r'on.?campus (?:recruit|interview|hiring)',
]


def is_academic_email(sender: str, subject: str, body: str) -> bool:
    sender_lower = sender.lower()
    is_edu_sender = bool(re.search(r'@[^>\s]*\.edu\b', sender_lower))
    combined = (subject + ' ' + body[:1000]).lower()

    if is_edu_sender:
        for pattern in JOB_OVERRIDE_PATTERNS:
            if re.search(pattern, combined):
                return False
        return True

    hits = sum(1 for p in ACADEMIC_CONTENT_PATTERNS if re.search(p, combined))
    return hits >= 2


# ── Regex status detection ────────────────────────────────────────────────────
NOISE_PATTERNS = [
    r'job alert',
    r'new jobs? (?:for|matching)',
    r'recommended jobs?',
    r'\d+ new jobs?',
    r'jobs? you might like',
    r'based on your (?:profile|resume|search)',
    r'(?:open|new) (?:roles?|positions?) (?:at|near)',
]

OFFER_PATTERNS = [
    r'pleased to (?:offer|extend an offer)',
    r"we(?:'re| are) (?:excited|pleased|happy|delighted) to offer you",
    r"we(?:'d| would) like to offer you",
    r'offer letter',
    r'congratulations.*(?:joining|new role|new position|accepted)',
    r'(?:joining|accepted).*congratulations',
    r'you have been selected.*(?:join|offer)',
    r'welcome (?:aboard|to the team)',
    r'compensation package',
    r'sign(?:ing)? bonus',
    r'your start date',
]

REJECTED_PATTERNS = [
    r'unfortunately.*(?:not|unable|decided|move)',
    r'not (?:moving|proceed)ing forward with your',
    r'moved forward with (?:other|another) candidate',
    r'not selected for (?:this|the)',
    r'decided to pursue other candidates',
    r'will not be moving forward',
    r'position has been filled',
    r'not (?:a match|the right fit) for',
    r'no longer (?:being considered|moving forward with you)',
    r"we(?:'re| have) decided not to",
    r"we've decided to move forward with (?:other|another)",
    r'after (?:careful )?consideration.*(?:not|decided)',
    r'wish you (?:all the best|success) in your (?:job )?search',
    r'we (?:will not|won\'t) be moving',
    r'we regret to inform',
]

INTERVIEW_PATTERNS = [
    r'calendly\.com',
    r'zoom\.us/[a-z]',
    r'meet\.google\.com',
    r'teams\.microsoft\.com/l/meetup',
    r'hackerrank\.com',
    r'codility\.com',
    r'codesignal\.com',
    r'hirevue\.com',
    r'take.?home (?:assignment|test|project|challenge)',
    r'technical (?:screen|round)\b',
    r'hiring manager (?:interview|call|round)',
    r'on.?site (?:interview|visit)',
    r'final (?:round|interview)',
    r'please (?:select|schedule|choose|pick|book) (?:a )?(?:time|slot|date)',
    r'schedule (?:a )?(?:30|45|60).?min',
    r'(?:phone|video) (?:screen|interview) (?:scheduled|confirmed)',
    r'interview (?:scheduled|confirmed|invitation)',
    r'we(?:\'d| would) like to (?:invite|schedule) you for (?:an )?interview',
]

APPLIED_PATTERNS = [
    r'(?:your )?application (?:has been |was )?(?:received|submitted|confirmed)',
    r'thank you for (?:applying|your application)',
    r'we(?:\'ve| have) received your application',
    r'application (?:successfully )?submitted',
    r'successfully applied',
    r'application confirmation',
]


def detect_status_regex(text: str) -> str:
    t = text.lower()

    for pattern in NOISE_PATTERNS:
        if re.search(pattern, t):
            return 'Unknown'

    for pattern in OFFER_PATTERNS:
        if re.search(pattern, t):
            return 'Offer'

    for pattern in REJECTED_PATTERNS:
        if re.search(pattern, t):
            return 'Rejected'

    for pattern in INTERVIEW_PATTERNS:
        if re.search(pattern, t):
            return 'Interview'

    for pattern in APPLIED_PATTERNS:
        if re.search(pattern, t):
            return 'Applied'

    return 'Unknown'


# ── Ollama classification ─────────────────────────────────────────────────────
_OLLAMA_PROMPT = """\
Classify this job application email into exactly one of these categories:
Offer, Rejected, Interview, Applied, Unknown

Definitions:
- Offer: Company explicitly offers you a job (mentions compensation, start date, "welcome aboard", signing bonus, or "we'd like to offer you")
- Rejected: Company explicitly declines your application ("not moving forward", "we regret to inform", "decided to pursue other candidates")
- Interview: Company wants to schedule an interview or assessment with YOU specifically (scheduling link, HackerRank/Codility test, "we'd like to invite you for an interview")
- Applied: Company confirms they received your application ("thank you for applying", "application received")
- Unknown: Job alert, marketing email, newsletter, or unclear

Email Subject: {subject}

Email Body:
{body}

Reply with one word only — the category name:"""


def classify_with_ollama(subject: str, body: str, model: str) -> str | None:
    """Returns classified status or None on failure (caller falls back to regex)."""
    prompt = _OLLAMA_PROMPT.format(subject=subject, body=body[:3000])
    try:
        resp = http.post(
            OLLAMA_URL,
            json={
                'model': model,
                'messages': [{'role': 'user', 'content': prompt}],
                'stream': False,
                'options': {'temperature': 0, 'num_predict': 8},
            },
            timeout=60,
        )
        resp.raise_for_status()
        raw = resp.json()['message']['content'].strip()
        word = raw.split()[0].strip('.,!?').capitalize()
        return word if word in VALID_STATUSES else 'Unknown'
    except Exception:
        return None


# ── Job board domains ────────────────────────────────────────────────────────
JOB_BOARDS = {
    'LinkedIn':   ['linkedin.com'],
    'Indeed':     ['indeed.com'],
    'Greenhouse': ['greenhouse.io'],
    'Lever':      ['lever.co'],
    'Workday':    ['myworkday.com', 'workday.com'],
    'Glassdoor':  ['glassdoor.com'],
}

# ── Role extraction ──────────────────────────────────────────────────────────
ROLE_PATTERNS = [
    r'(?:position|role|job|opportunity) (?:of |for |as )?(?:a |an )?([A-Za-z][A-Za-z\s/,-]+?)(?:\s+(?:at|with|in)|[,.]|$)',
    r'(?:applying|applied) (?:for |to )?(?:the )?([A-Za-z][A-Za-z\s/,-]+?)(?:\s+(?:position|role|job)|[,.]|$)',
    r'([A-Za-z][A-Za-z\s/,-]+?) (?:position|role|opening|opportunity)\b',
    r'(?:re:|fw:)\s*(?:your application for |application - )?([A-Za-z][A-Za-z\s/,-]+?)(?:\s+at|\s*[-|]|\s*$)',
]


# ── Auth ─────────────────────────────────────────────────────────────────────
def get_gmail_service():
    creds = None
    if TOKEN_FILE.exists():
        with open(TOKEN_FILE) as f:
            creds = Credentials.from_authorized_user_info(json.load(f), SCOPES)

    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(str(CREDS_FILE), SCOPES)
            creds = flow.run_local_server(port=0)
        with open(TOKEN_FILE, 'w') as f:
            f.write(creds.to_json())

    return build('gmail', 'v1', credentials=creds)


# ── Body extraction ───────────────────────────────────────────────────────────
def extract_body(payload: dict) -> str:
    mime = payload.get('mimeType', '')
    if mime == 'text/plain':
        data = payload.get('body', {}).get('data', '')
        if data:
            return base64.urlsafe_b64decode(data + '==').decode('utf-8', errors='ignore')

    for part in payload.get('parts', []):
        result = extract_body(part)
        if result:
            return result

    return ''


def extract_company(sender: str) -> str:
    m = re.match(r'"?([^"<@\n]+?)"?\s*<', sender)
    if m:
        name = m.group(1).strip()
        skip = {'noreply', 'no-reply', 'recruiting', 'talent', 'careers',
                 'hr', 'jobs', 'hiring', 'notifications', 'info', 'hello',
                 'team', 'support', 'do not reply'}
        if name.lower() not in skip and len(name) > 1:
            return name

    m = re.search(r'@([^.@>]+)\.', sender)
    if m:
        domain = m.group(1)
        skip_domains = {'gmail', 'yahoo', 'hotmail', 'outlook', 'noreply',
                        'no-reply', 'mail', 'email', 'bounce', 'send'}
        if domain.lower() not in skip_domains:
            return domain.capitalize()

    return 'Unknown'


def extract_role(subject: str, body: str) -> str:
    clean_subject = re.sub(r'^(re|fw|fwd):\s*', '', subject, flags=re.IGNORECASE).strip()
    text = clean_subject + ' ' + body[:500]

    for pattern in ROLE_PATTERNS:
        m = re.search(pattern, text, re.IGNORECASE)
        if m:
            role = m.group(1).strip().rstrip('.,')
            if 3 < len(role) < 70:
                return role

    return 'Unknown'


def detect_source(sender: str) -> str:
    sender_lower = sender.lower()
    for board, domains in JOB_BOARDS.items():
        if any(d in sender_lower for d in domains):
            return board
    return 'Direct'


def parse_message(msg: dict, use_ml: bool = False, ollama_model: str = OLLAMA_MODEL_DEFAULT) -> dict | None:
    headers = {h['name']: h['value'] for h in msg['payload']['headers']}
    subject = headers.get('Subject', '')
    sender  = headers.get('From', '')
    snippet = msg.get('snippet', '')
    body    = extract_body(msg['payload'])
    text    = body if body else snippet

    if is_blocked(sender, subject):
        return None

    if is_academic_email(sender, subject, text):
        return None

    try:
        ts   = int(msg.get('internalDate', 0))
        date = datetime.fromtimestamp(ts / 1000).strftime('%Y-%m-%d')
    except Exception:
        date = ''

    detection_text = f"{subject} {text}"
    regex_status   = detect_status_regex(detection_text)

    if use_ml and regex_status == 'Unknown':
        ml_status = classify_with_ollama(subject, text, ollama_model)
        status    = ml_status if ml_status is not None else regex_status
    else:
        status = regex_status

    record = {
        'Date':         date,
        'Company':      extract_company(sender),
        'Role':         extract_role(subject, text),
        'Status':       status,
        'Regex_Status': regex_status,
        'Source':       detect_source(sender),
        'Subject':      subject,
        'Sender':       sender,
        'Snippet':      snippet[:250],
        'Message ID':   msg['id'],
        'Thread ID':    msg['threadId'],
    }

    if not use_ml:
        del record['Regex_Status']

    return record


# ── Fetch ────────────────────────────────────────────────────────────────────
def fetch_message_ids(service, max_results: int) -> list[dict]:
    print(f"Searching Gmail (up to {max_results} results)...")
    refs, page_token = [], None

    while len(refs) < max_results:
        batch = min(100, max_results - len(refs))
        kwargs = {'userId': 'me', 'q': GMAIL_QUERY, 'maxResults': batch}
        if page_token:
            kwargs['pageToken'] = page_token

        resp       = service.users().messages().list(**kwargs).execute()
        refs      += resp.get('messages', [])
        page_token = resp.get('nextPageToken')
        if not page_token:
            break

    print(f"Found {len(refs)} matching emails.")
    return refs


# ── Excel export with colour-coding + auto-fit ───────────────────────────────
_STATUS_FILLS = {
    'Offer':     PatternFill('solid', fgColor='C6EFCE'),  # green
    'Interview': PatternFill('solid', fgColor='FFEB9C'),  # yellow
    'Applied':   PatternFill('solid', fgColor='BDD7EE'),  # blue
    'Rejected':  PatternFill('solid', fgColor='FFC7CE'),  # red
    'Unknown':   PatternFill('solid', fgColor='EDEDED'),  # grey
}
_STATUS_FONTS = {
    'Offer':     Font(bold=True, color='006100'),
    'Interview': Font(bold=True, color='9C6500'),
    'Applied':   Font(bold=True, color='1F4E79'),
    'Rejected':  Font(bold=True, color='9C0006'),
    'Unknown':   Font(color='595959'),
}
_HEADER_FILL = PatternFill('solid', fgColor='2F5496')
_HEADER_FONT = Font(bold=True, color='FFFFFF')
_THIN_BORDER = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9'),
)


def save_excel(df: 'pd.DataFrame', path: 'Path') -> None:
    df.to_excel(path, index=False, engine='openpyxl')

    from openpyxl import load_workbook
    wb = load_workbook(path)
    ws = wb.active

    # Style header row
    for cell in ws[1]:
        cell.fill      = _HEADER_FILL
        cell.font      = _HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)
        cell.border    = _THIN_BORDER

    # Find Status column index (1-based)
    status_col = None
    for idx, cell in enumerate(ws[1], 1):
        if cell.value == 'Status':
            status_col = idx
            break

    # Style data rows
    for row in ws.iter_rows(min_row=2):
        status_val = row[status_col - 1].value if status_col else None
        fill = _STATUS_FILLS.get(status_val)
        font = _STATUS_FONTS.get(status_val)
        for cell in row:
            cell.border    = _THIN_BORDER
            cell.alignment = Alignment(vertical='center', wrap_text=False)
            if cell.column == status_col:
                if fill:
                    cell.fill = fill
                if font:
                    cell.font = font

    # Auto-fit column widths
    for col_cells in ws.columns:
        max_len = max(
            (len(str(c.value)) if c.value is not None else 0) for c in col_cells
        )
        col_letter = get_column_letter(col_cells[0].column)
        ws.column_dimensions[col_letter].width = min(max_len + 4, 60)

    # Freeze top row
    ws.freeze_panes = 'A2'

    wb.save(path)


# ── Main ─────────────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description='Gmail Job Tracker')
    parser.add_argument('--ml', action='store_true',
                        help='Use Ollama to classify emails regex marks Unknown')
    parser.add_argument('--model', default=OLLAMA_MODEL_DEFAULT,
                        help=f'Ollama model to use (default: {OLLAMA_MODEL_DEFAULT})')
    args = parser.parse_args()

    if args.ml:
        # Verify Ollama is reachable before starting
        try:
            r = http.get('http://localhost:11434/api/tags', timeout=5)
            r.raise_for_status()
            print(f"Ollama ready. Model: {args.model}")
        except Exception:
            print("[!] Ollama not reachable at localhost:11434. Run: ollama serve")
            print("    Falling back to regex-only mode.")
            args.ml = False

    service = get_gmail_service()

    if OUTPUT_FILE.exists():
        existing_df = pd.read_excel(OUTPUT_FILE, engine='openpyxl')
        seen_ids    = set(existing_df['Message ID'].astype(str))
        print(f"Existing records: {len(existing_df)}")
    else:
        existing_df = pd.DataFrame()
        seen_ids    = set()

    refs     = fetch_message_ids(service, MAX_RESULTS)
    new_refs = [r for r in refs if r['id'] not in seen_ids]
    print(f"New emails to process: {len(new_refs)}")

    records = []
    skipped = 0
    ml_used = 0

    for i, ref in enumerate(new_refs, 1):
        try:
            msg = service.users().messages().get(
                userId='me', id=ref['id'], format='full',
            ).execute()
            record = parse_message(msg, use_ml=args.ml, ollama_model=args.model)
            if record is None:
                skipped += 1
            else:
                if args.ml and record.get('Regex_Status') == 'Unknown' and record['Status'] != 'Unknown':
                    ml_used += 1
                records.append(record)
        except Exception as e:
            print(f"  [!] Error on {ref['id']}: {e}")

        if i % 50 == 0:
            print(f"  Processed {i}/{len(new_refs)}...")

    print(f"Skipped (academic/university): {skipped}")
    if args.ml:
        print(f"Ollama reclassified from Unknown: {ml_used}")

    if records:
        new_df   = pd.DataFrame(records)
        final_df = pd.concat([existing_df, new_df], ignore_index=True)
        final_df.sort_values('Date', ascending=False, inplace=True)
        save_excel(final_df, OUTPUT_FILE)
        print(f"\nSaved {len(final_df)} total records → {OUTPUT_FILE}")
    else:
        print("\nNo new records — sheet unchanged.")
        final_df = existing_df

    if not final_df.empty:
        print("\n─── Status Summary ───────────────────")
        print(final_df['Status'].value_counts().to_string())
        print("──────────────────────────────────────")


if __name__ == '__main__':
    main()
